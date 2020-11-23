import openpyxl

# row count
def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

# column count
def getColumn(file, sheetName):
    worksheet = openpyxl.load_workbook(file)
    sheet = worksheet[sheetName]
    return (sheet.max_column)

# read data from excel
def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row = rownum, column = columnno).value

# write data into excel
def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row = rownum, column = columnno).value = data
    workbook.save(file)
